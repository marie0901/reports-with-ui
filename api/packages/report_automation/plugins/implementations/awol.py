"""AWOL Plugin - Inactive users campaigns report."""

import pandas as pd
from pathlib import Path
from typing import Dict, List
import logging
from openpyxl import Workbook, load_workbook
import copy

from ..base import BaseReportPlugin, register_plugin
from ..constants import WEEKLY_BOUNDARIES

logger = logging.getLogger(__name__)

AWOL_MAPPINGS = {"Day 1": "1d", "Day 3": "3d", "Day 5": "5d", "Day 10": "10d", "Day 15": "15d", "Day 20": "20d", "Day 30": "30d", "Day 40": "40d"}
METRICS = ["sent", "delivered", "opened", "clicked", "converted", "unsubscribed"]
WEEK_COLUMNS = {
    'week1': 'Q', 'week2': 'P', 'week3': 'O', 'week4': 'N', 'week5': 'M', 'week6': 'L', 'week7': 'K', 'week8': 'J',
    'week9': 'I', 'week10': 'H', 'week11': 'G', 'week12': 'F'
}
TIMING_BLOCKS = {
    "1d": {"inactive7": [3, 10], "inactive14": [11, 18], "inactive22": [35, 42], "inactive31": [59, 66]},
    "3d": {"inactive14": [19, 26], "inactive22": [43, 50], "inactive31": [67, 74]},
    "5d": {"inactive22": [43, 50], "inactive31": [67, 74]},
    "10d": {"inactive31": [75, 82]},
    "15d": {"inactive31": [83, 90]},
    "20d": {"inactive31": [91, 98]},
    "30d": {"inactive31": [99, 106]},
    "40d": {"inactive31": [107, 114]},
}
WEEK_MAPPINGS = {
    'source': {
        '01': 'Q', '02': 'P', '03': 'O', '04': 'N', '05': 'M', '06': 'L', '07': 'K', '08': 'J',
        '09': 'I', '10': 'H', '11': 'G', '12': 'F'
    },
    'target': {
        '01': 'BF', '02': 'BE', '03': 'BD', '04': 'BC', '05': 'BB', '06': 'BA', '07': 'AZ', '08': 'AY',
        '09': 'AX', '10': 'AW', '11': 'AV', '12': 'AU'
    }
}

SHEET_MAPPINGS = {
    'awol': 'AWOL Chains Sport'
}


@register_plugin
class AWOLPlugin(BaseReportPlugin):
    """AWOL report plugin."""
    
    name = "awol"
    supports_multiple_files = True
    
    def __init__(self):
        self.existing_excel = None
        self.replace_week = None
    
    def process_csv(self, csv_paths: List[Path]) -> Dict[str, pd.DataFrame]:
        data_files = {}
        for path in csv_paths:
            df = pd.read_csv(path)
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
            data_files[path.name] = df
            logger.info(f"Loaded {path.name}: {len(df)} rows")
        return data_files
    
    def transform_data(self, data_files: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, pd.DataFrame]]:
        report_data = {}
        
        for file_name, data in data_files.items():
            # Detect campaign from data
            campaign_key = None
            if 'campaign_name' in data.columns and not data.empty:
                campaign_name = data['campaign_name'].iloc[0]
                logger.info(f"File: {file_name}, Campaign name from CSV: '{campaign_name}'")
                campaign_name_lower = campaign_name.lower()
                # Extract number from campaign name (e.g., "Inactive 7" -> 7, "Inactive 31+" -> 31)
                import re
                match = re.search(r'inactive\s*(\d+)', campaign_name_lower)
                if match:
                    num = int(match.group(1))
                    logger.info(f"Extracted number: {num}")
                    if num == 7:
                        campaign_key = 'inactive7'
                    elif num == 14:
                        campaign_key = 'inactive14'
                    elif num in [21, 22]:
                        campaign_key = 'inactive22'
                    elif num >= 30:
                        campaign_key = 'inactive31'
                    logger.info(f"Mapped to campaign_key: {campaign_key}")
                else:
                    logger.warning(f"No 'inactive' pattern found in campaign name: '{campaign_name}'")
            
            # Log available templates
            if 'template_name' in data.columns:
                unique_templates = data['template_name'].unique()
                logger.info(f"Templates in {file_name}: {list(unique_templates)}")
            
            # Log date range in the data
            if 'datetime' in data.columns and not data.empty:
                min_date = data['datetime'].min()
                max_date = data['datetime'].max()
                logger.info(f"Date range in {file_name}: {min_date} to {max_date}")
            
            filtered = data[data['template_name'].isin(AWOL_MAPPINGS.keys())]
            logger.info(f"Filtered {len(filtered)} rows matching AWOL_MAPPINGS from {len(data)} total rows")
            if not filtered.empty:
                logger.info(f"Filtered date range: {filtered['datetime'].min()} to {filtered['datetime'].max()}")
            
            weekly_data = {}
            for i, (start_date, end_date) in enumerate(WEEKLY_BOUNDARIES, 1):
                start = pd.to_datetime(start_date + ' 00:00:00')
                end = pd.to_datetime(end_date + ' 23:59:59')
                week_mask = (filtered['datetime'] >= start) & (filtered['datetime'] <= end)
                week_df = filtered[week_mask].groupby('template_name')[METRICS].sum().reset_index()
                weekly_data[f'week{i}'] = week_df
                if i == 9 and not week_df.empty:
                    logger.info(f"Week 9 data for {file_name}: {len(week_df)} templates with data")
                    logger.info(f"Week 9 date range: {start_date} to {end_date}")
                    for _, row in week_df.iterrows():
                        logger.info(f"  Template '{row['template_name']}': sent={row['sent']}, delivered={row['delivered']}")
            
            file_report = {}
            for template_name, timing_category in AWOL_MAPPINGS.items():
                timing_data = {}
                for week_key, week_df in weekly_data.items():
                    template_data = week_df[week_df['template_name'] == template_name]
                    if not template_data.empty:
                        template_data = self._calculate_percentages(template_data)
                        timing_data[week_key] = template_data
                    else:
                        timing_data[week_key] = self._empty_data(template_name)
                file_report[timing_category] = timing_data
            
            report_data[file_name] = {'campaign_key': campaign_key, 'data': file_report}
        
        return report_data
    
    def _calculate_percentages(self, data: pd.DataFrame) -> pd.DataFrame:
        result = data.copy()
        result['pct_delivered'] = (result['delivered'] / result['sent'] * 100).fillna(0)
        result['pct_open'] = (result['opened'] / result['delivered'] * 100).fillna(0)
        result['pct_click'] = (result['clicked'] / result['delivered'] * 100).fillna(0)
        result['pct_cr'] = (result['converted'] / result['delivered'] * 100).fillna(0)
        return result
    
    def _empty_data(self, template_name: str) -> pd.DataFrame:
        return pd.DataFrame([{
            'template_name': template_name, 'sent': 0, 'delivered': 0, 'opened': 0,
            'clicked': 0, 'converted': 0, 'unsubscribed': 0,
            'pct_delivered': 0, 'pct_open': 0, 'pct_click': 0, 'pct_cr': 0
        }])
    
    def generate_excel(self, report_data: Dict[str, Dict[str, pd.DataFrame]], output_path: Path):
        wb = Workbook()
        ws = wb.active
        
        week_headers = {
            'week8': ("08", "16.02"), 'week7': ("07", "09.02"),
            'week6': ("06", "02.02"), 'week5': ("05", "26.01"), 'week4': ("04", "19.01"),
            'week3': ("03", "12.01"), 'week2': ("02", "05.01"), 'week1': ("01", "29.12")
        }
        for week_key, col_letter in WEEK_COLUMNS.items():
            if week_key in week_headers:
                week_display, date = week_headers[week_key]
                ws[f'{col_letter}1'] = f"Week {week_display}\n{date}"
        
        logger.info(f"Processing {len(report_data)} files")
        for file_name, file_info in report_data.items():
            campaign_key = file_info.get('campaign_key')
            section_data = file_info.get('data')
            
            logger.info(f"File: {file_name}, Campaign: {campaign_key}")
            
            if campaign_key == 'inactive7':
                self._populate_section(ws, section_data, "inactive7", "Inactive 7 [SPORT] ⚽️")
            elif campaign_key == 'inactive14':
                self._populate_section(ws, section_data, "inactive14", "Inactive 14 [SPORT] ⚽️")
            elif campaign_key == 'inactive22':
                self._populate_section(ws, section_data, "inactive22", "Inactive 22 [SPORT] ⚽️")
            elif campaign_key == 'inactive31':
                self._populate_section(ws, section_data, "inactive31", "Inactive 31+ [SPORT] ⚽️")
        
        wb.save(output_path)
        logger.info(f"Excel saved: {output_path}")
        
        if self.existing_excel and self.replace_week:
            self._replace_week(output_path, self.existing_excel, self.replace_week)
    
    def _populate_section(self, ws, section_data: Dict, section_key: str, campaign_name: str):
        logger.info(f"Populating section: {section_key}, campaign: {campaign_name}")
        logger.info(f"Available timing categories in data: {list(section_data.keys())}")
        
        templates_with_data = []
        for timing_category in section_data.keys():
            has_data = any(not df.empty and df['sent'].iloc[0] > 0 
                          for df in section_data[timing_category].values() 
                          if not df.empty)
            if has_data:
                templates_with_data.append(timing_category)
        
        templates_with_data.sort(key=lambda x: int(x.replace('d', '')))
        logger.info(f"Templates to populate (sorted): {templates_with_data}")
        
        for timing_category in templates_with_data:
            if timing_category not in TIMING_BLOCKS:
                logger.warning(f"Timing category '{timing_category}' not in TIMING_BLOCKS, skipping")
                continue
            
            if section_key not in TIMING_BLOCKS[timing_category]:
                logger.warning(f"Section '{section_key}' not in TIMING_BLOCKS['{timing_category}'], skipping")
                continue
            
            start_row, end_row = TIMING_BLOCKS[timing_category][section_key]
            template_name = {v: k for k, v in AWOL_MAPPINGS.items()}.get(timing_category, timing_category)
            logger.info(f"Writing {timing_category} ({template_name}) to rows {start_row}-{end_row}")
            
            ws[f'B{start_row}'] = campaign_name.replace(" [SPORT] ⚽️", "").lower()
            ws[f'C{start_row}'] = " All Mail"
            ws[f'D{start_row}'] = template_name
            
            timing_data = section_data[timing_category]
            metrics = ["sent", "delivered", "opened", "clicked", "unsubscribed", "pct_delivered", "pct_open", "pct_click"]
            
            for i, metric in enumerate(metrics):
                row = start_row + i
                
                for week_key, col_letter in WEEK_COLUMNS.items():
                    if week_key in timing_data and not timing_data[week_key].empty:
                        df = timing_data[week_key]
                        if metric in df.columns:
                            ws[f'{col_letter}{row}'] = df[metric].iloc[0]
                    else:
                        ws[f'{col_letter}{row}'] = 0
    
    def _replace_week(self, generated_path: Path, existing_path: Path, week_number: str):
        source_col = WEEK_MAPPINGS['source'][week_number]
        target_col = WEEK_MAPPINGS['target'][week_number]
        
        logger.info(f"Starting week replacement: week {week_number}, source col {source_col}, target col {target_col}")
        
        generated_wb = load_workbook(generated_path, data_only=False)
        existing_wb = load_workbook(existing_path, data_only=False)
        
        generated_ws = generated_wb.active
        
        # Get target sheet based on report type
        sheet_name = SHEET_MAPPINGS.get('awol', existing_wb.sheetnames[0])
        logger.info(f"Target sheet name: '{sheet_name}'")
        logger.info(f"Available sheets in workbook: {existing_wb.sheetnames}")
        if sheet_name in existing_wb.sheetnames:
            existing_ws = existing_wb[sheet_name]
            logger.info(f"Using sheet: '{sheet_name}'")
        else:
            existing_ws = existing_wb.active
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet: '{existing_ws.title}'")
        
        self._copy_formatting(existing_ws, 'BE', target_col)
        
        copied = 0
        
        source_blocks = {}
        for source_row in range(1, generated_ws.max_row + 1):
            source_b = generated_ws[f'B{source_row}'].value
            source_d = generated_ws[f'D{source_row}'].value
            if source_b and source_d:
                key = (str(source_b).strip().lower(), str(source_d).strip())
                source_blocks[key] = source_row
        
        logger.info(f"Found {len(source_blocks)} source blocks: {list(source_blocks.keys())}")
        
        expected_blocks = [
            ("inactive 7", "Day 1"),
            ("inactive 14", "Day 1"), ("inactive 14", "Day 3"),
            ("inactive 22", "Day 1"), ("inactive 22", "Day 5"),
            ("inactive 31+", "Day 1"), ("inactive 31+", "Day 5"), ("inactive 31+", "Day 10"),
            ("inactive 31+", "Day 15"), ("inactive 31+", "Day 20"), ("inactive 31+", "Day 30"), ("inactive 31+", "Day 40")
        ]
        missing = [b for b in expected_blocks if b not in source_blocks]
        if missing:
            logger.warning(f"Missing blocks in source: {missing}")
        
        for (campaign, template), source_start in source_blocks.items():
            logger.info(f"Looking for campaign='{campaign}', template='{template}' in target Excel")
            target_start = self._find_target_block(existing_ws, campaign, template)
            if target_start:
                logger.info(f"Found target at row {target_start}")
                skipped_formulas = 0
                for offset in range(8):  # 8 metrics per block
                    source_row = source_start + offset
                    target_row = target_start + offset
                    
                    # Skip if target cell has a formula
                    target_cell = existing_ws[f'{target_col}{target_row}']
                    if isinstance(target_cell.value, str) and target_cell.value.startswith('='):
                        skipped_formulas += 1
                        continue
                    
                    value = generated_ws[f'{source_col}{source_row}'].value
                    if value is not None:
                        existing_ws[f'{target_col}{target_row}'].value = value
                        copied += 1
                    elif offset < 5:
                        logger.debug(f"No value at source row {source_row} for metric offset {offset}")
                
                if skipped_formulas > 0:
                    logger.info(f"Skipped {skipped_formulas} formula cells for {campaign}/{template}")
            else:
                logger.warning(f"Target not found for campaign='{campaign}', template='{template}'")
        
        # Save the updated existing workbook back to the generated_path
        # This ensures the API returns the updated file instead of the intermediate report
        existing_wb.save(generated_path)
        logger.info(f"Updated Excel saved to output path: {generated_path} ({copied} values)")
    
    def _copy_formatting(self, ws, source_col: str, target_col: str):
        for row in range(1, ws.max_row + 1):
            source_cell = ws[f'{source_col}{row}']
            target_cell = ws[f'{target_col}{row}']
            # Only copy style, not formulas
            if source_cell.has_style and not (isinstance(source_cell.value, str) and source_cell.value and source_cell.value.startswith('=')):
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
    
    def _find_target_block(self, ws, campaign: str, template: str) -> int:
        # First find campaign start
        campaign_start = None
        logger.debug(f"Searching for campaign '{campaign}' in column B")
        for row in range(1, ws.max_row + 1):
            cell_b = ws[f'B{row}'].value
            if cell_b and campaign in str(cell_b).lower():
                campaign_start = row
                logger.debug(f"Found campaign at row {row}: '{cell_b}'")
                break
        
        if not campaign_start:
            logger.warning(f"Campaign '{campaign}' not found in target Excel")
            return None
        
        # Then find template within campaign section (next 100 rows)
        logger.debug(f"Searching for template '{template}' starting from row {campaign_start}")
        for row in range(campaign_start, min(campaign_start + 100, ws.max_row + 1)):
            cell_d = ws[f'D{row}'].value
            if cell_d:
                logger.debug(f"Row {row}, column D: '{cell_d}'")
                if template.lower() in str(cell_d).lower():
                    logger.debug(f"Found template match at row {row}")
                    return row
            # Stop if we hit next campaign
            cell_b = ws[f'B{row}'].value
            if cell_b and row > campaign_start and str(cell_b).strip():
                logger.debug(f"Hit next campaign at row {row}, stopping search")
                break
        
        logger.warning(f"Template '{template}' not found for campaign '{campaign}'")
        return None
    
    def execute(self, input_paths: List[Path], output_path: Path, existing_excel: Path = None, replace_week: str = None):
        self.existing_excel = existing_excel
        self.replace_week = replace_week
        
        for path in input_paths:
            self.validate_input(path)
        
        data_files = self.process_csv(input_paths)
        report_data = self.transform_data(data_files)
        self.generate_excel(report_data, output_path)
