"""Dynamic plugin that uses JSON configuration for report generation."""

import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
import copy

from ..base import BaseReportPlugin
from ..constants import WEEKLY_BOUNDARIES

import re
import unicodedata

logger = logging.getLogger(__name__)

def normalize_name(name: str) -> str:
    """Normalize names for better matching (lowercase, no emojis/symbols)."""
    if not name: return ""
    # Remove emojis and special symbols
    name = "".join(ch for ch in name if unicodedata.category(ch)[0] != 'S')
    # Lowercase, strip and remove extra spaces
    return " ".join(name.lower().split())

class DynamicReportPlugin(BaseReportPlugin):
    """A report plugin that is driven by a JSON configuration."""
    
    def __init__(self, config_path: Path):
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        
        self._name = self.config.get('id', config_path.stem)
        self.existing_excel = None
        self.replace_week = None
        self._original_files = []

    @property
    def name(self) -> str:
        return self._name

    @property
    def supports_multiple_files(self) -> bool:
        # Most of our reports require multiple files or at least support them
        return True

    def process_csv(self, csv_paths: List[Path]) -> Dict[str, pd.DataFrame]:
        data_files = {}
        for path in csv_paths:
            try:
                df = pd.read_csv(path)
                df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
                data_files[path.name] = df
                logger.info(f"Loaded {path.name}: {len(df)} rows")
            except Exception as e:
                logger.error(f"Error loading {path.name}: {e}")
        return data_files

    def _detect_section_by_campaign(self, data: pd.DataFrame) -> Optional[Dict[str, Any]]:
        """Detect section based on campaign mapping in config."""
        if data.empty or 'campaign_name' not in data.columns:
            return None
        
        campaign_name = data['campaign_name'].iloc[0]
        norm_campaign = normalize_name(campaign_name)
        
        logger.info(f"Detecting section for campaign: '{campaign_name}' (norm: '{norm_campaign}')")
        
        for mapping in self.config.get('campaign_mappings', []):
            mapped_name = mapping['campaign_name']
            norm_mapped = normalize_name(mapped_name)
            
            # Exact match first
            if mapped_name == campaign_name:
                logger.debug(f"Matches exact mapping: {mapped_name}")
                return mapping
                
            # Normalized match
            if norm_mapped == norm_campaign:
                logger.debug(f"Matches normalized mapping: {mapped_name}")
                return mapping
                
            # Fuzzy/contained match
            if (mapping.get('fuzzy_match') or True) and (norm_mapped in norm_campaign or norm_campaign in norm_mapped):
                logger.info(f"Matches fuzzy mapping: {mapped_name}")
                return mapping
                
        logger.warning(f"No mapping found for campaign: {campaign_name}")
        return None

    def transform_data(self, data_files: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Transform data based on template mappings."""
        report_data = {}
        
        # Flatten template mappings for easy lookup
        template_lookup = {}
        for campaign_mapping in self.config.get('template_mappings', []):
            campaign = campaign_mapping['campaign_name']
            for t in campaign_mapping.get('templates', []):
                template_lookup[(campaign, t['csv_template_name'])] = t
        
        for file_name, data in data_files.items():
            campaign_info = self._detect_section_by_campaign(data)
            if not campaign_info:
                logger.warning(f"Skipping {file_name}: unknown campaign")
                continue
            
            campaign_name = campaign_info['campaign_name']
            
            # Filter templates relevant to this campaign
            relevant_templates = [t['csv_template_name'] for (c, t_name), t in template_lookup.items() if c == campaign_name]
            filtered = data[data['template_name'].isin(relevant_templates)]
            
            # Aggregate by weeks
            weekly_data = {}
            # Dynamically determine metrics to aggregate based on config
            metrics_structure = self.config.get('metrics', {}).get('row_structure', [])
            data_metrics = [m['name'].lower().replace(' ', '_') for m in metrics_structure if m['type'] == 'data']
            # Fallback to defaults if none found
            if not data_metrics:
                data_metrics = ['sent', 'delivered', 'opened', 'clicked', 'unsubscribed']
            
            # Ensure columns exist in dataframe
            data_metrics = [m for m in data_metrics if m in filtered.columns]
            
            logger.info(f"Aggregating {len(filtered)} rows with metrics: {data_metrics}")

            for i, (start_date, end_date) in enumerate(WEEKLY_BOUNDARIES, 1):
                start = pd.to_datetime(start_date + ' 00:00:00')
                end = pd.to_datetime(end_date + ' 23:59:59')
                week_mask = (filtered['datetime'] >= start) & (filtered['datetime'] <= end)
                week_df = filtered[week_mask].groupby('template_name')[data_metrics].sum().reset_index()
                weekly_data[f'week{i:02d}'] = week_df # standardized 01, 02 format
                weekly_data[f'week{i}'] = week_df # also support 1, 2 format
            
            file_report = {}
            for t_name in relevant_templates:
                t_info = template_lookup[(campaign_name, t_name)]
                timing_data = {}
                for week_key, week_df in weekly_data.items():
                    t_df = week_df[week_df['template_name'] == t_name]
                    if not t_df.empty:
                        timing_data[week_key] = t_df.iloc[0].to_dict()
                    else:
                        timing_data[week_key] = {'sent': 0, 'delivered': 0, 'opened': 0, 'clicked': 0, 'unsubscribed': 0}
                file_report[t_name] = timing_data
            
            report_data[file_name] = {
                'campaign': campaign_info,
                'templates': file_report
            }
            
        return report_data

    def generate_excel(self, report_data: Dict[str, Any], output_path: Path):
        """Standard generate_excel is usually for intermediate reports, 
        but our dynamic plugin mostly targets 'replace_week' functionality."""
        # For now, create a simple intermediate report if no existing_excel
        if not self.existing_excel:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Dynamic Report Intermediate Output"
            wb.save(output_path)
            return

        if self.replace_week:
            self._replace_week(output_path, self.existing_excel, self.replace_week, report_data)

    def _replace_week(self, generated_path: Path, existing_path: Path, week_number: str, report_data: Dict[str, Any]):
        """The core logic for writing to the target Excel file."""
        week_key = week_number.zfill(2)
        internal_week_key = f"week{week_key}" # The transform_data uses "week07" etc as keys
        
        target_col = self.config.get('week_columns', {}).get('mappings', {}).get(week_key)
        
        if not target_col:
            logger.error(f"No column mapping for week {week_number} (searched for key '{week_key}')")
            logger.info(f"Available mappings: {list(self.config.get('week_columns', {}).get('mappings', {}).keys())}")
            return

        logger.info(f"Starting replacement for {internal_week_key} (column {target_col})")

        existing_wb = load_workbook(existing_path, data_only=False)
        copied = 0
        
        # Process each file's worth of data
        for file_name, data in report_data.items():
            campaign_info = data['campaign']
            # Priority: Root config target_sheet -> Campaign mapping target_sheet -> first sheet
            sheet_name = self.config.get('target_sheet') or campaign_info.get('target_sheet')
            
            if not sheet_name:
                sheet_name = existing_wb.sheetnames[0]
                logger.info(f"No target sheet defined, using first sheet: {sheet_name}")

            if sheet_name not in existing_wb.sheetnames:
                # Try a fuzzy match for sheet name
                matches = [s for s in existing_wb.sheetnames if sheet_name.lower() in s.lower() or s.lower() in sheet_name.lower()]
                if matches:
                    logger.info(f"Sheet '{sheet_name}' not found, using fuzzy match: '{matches[0]}'")
                    sheet_name = matches[0]
                else:
                    logger.warning(f"Sheet {sheet_name} not found in Excel. Available: {existing_wb.sheetnames}")
                    continue
                
            ws = existing_wb[sheet_name]
            logger.info(f"Processing data from {file_name} into sheet '{sheet_name}'")
            
            # Template mappings for this campaign
            templates = data['templates']
            
            # Find template rows
            for t_name, week_data in templates.items():
                logger.debug(f"Checking template: {t_name}")
                # We need the excel_row from the config
                template_config = None
                for c_mapping in self.config.get('template_mappings', []):
                    if c_mapping['campaign_name'] == campaign_info['campaign_name']:
                        for t in c_mapping['templates']:
                            if t['csv_template_name'] == t_name:
                                template_config = t
                                break
                
                if not template_config:
                    logger.debug(f"No template mapping found in config for {t_name} in campaign {campaign_info['campaign_name']}")
                    continue
                    
                target_row = template_config['excel_row']
                metrics_structure = self.config.get('metrics', {}).get('row_structure', [])
                
                # Get data for THIS specific week
                specific_week_data = week_data.get(internal_week_key)
                if not specific_week_data:
                    logger.debug(f"No data for {t_name} in {internal_week_key}")
                    continue

                for metric_info in metrics_structure:
                    if metric_info['type'] == 'data':
                        metric_name = metric_info['name'].lower().replace(' ', '_')
                        val = specific_week_data.get(metric_name)
                        
                        if val is not None:
                            target_cell_row = target_row + metric_info['offset']
                            col_idx = self._col_to_num(target_col)
                            ws.cell(row=target_cell_row, column=col_idx, value=val)
                            copied += 1
                            logger.info(f"Copied {t_name} {metric_name}: {val} -> Row {target_cell_row} Col {target_col}")
                        else:
                            logger.debug(f"Metric {metric_name} not found in data for {t_name}")

        existing_wb.save(generated_path)
        logger.info(f"Updated dynamic report saved with {copied} values copied")

    def _col_to_num(self, col: str) -> int:
        num = 0
        for char in col:
            num = num * 26 + (ord(char.upper()) - ord('A') + 1)
        return num

    def execute(self, input_paths: List[Path], output_path: Path, existing_excel: Path = None, replace_week: str = None):
        self.existing_excel = existing_excel
        self.replace_week = replace_week
        self._original_files = input_paths
        
        data_files = self.process_csv(input_paths)
        report_data = self.transform_data(data_files)
        self.generate_excel(report_data, output_path)

    def get_input_slots(self) -> List[Dict[str, Any]]:
        return self.config.get('input_slots', super().get_input_slots())

    def get_expected_templates(self) -> Dict[str, List[str]]:
        # Map slot ID to expected template names
        expected = {}
        for campaign_mapping in self.config.get('template_mappings', []):
            templates = [t['csv_template_name'] for t in campaign_mapping.get('templates', [])]
            # For simplicity, if not specified, put all in 'source'
            expected['source'] = expected.get('source', []) + templates
        return expected
