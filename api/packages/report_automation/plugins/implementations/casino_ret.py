"""Casino-Ret Plugin - Casino and retention campaigns report."""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import copy

from ..base import BaseReportPlugin, register_plugin
from ..constants import WEEKLY_BOUNDARIES

logger = logging.getLogger(__name__)

# Campaign-based section detection mappings
CAMPAIGN_SECTION_MAPPINGS = {
    "casino+sport A/B Reg_No_Dep": {"section": "casino", "start_row": 3, "label": "casino+sport A/B Reg_No_Dep"},
    "Ret 1 dep [SPORT] ⚽️": {"section": "retention", "start_row": 75, "label": "Ret 1 dep [SPORT] ⚽️"},
    "Ret 2 dep [SPORT] ⚽️": {"section": "retention", "start_row": 123, "label": "Ret 2 dep [SPORT] ⚽️"},
}

# Template mappings
RETENTION_MAPPINGS = {"Day 3": "3d", "Day 4": "4d", "Day 6": "6d", "Day 8": "8d", "Day 10": "10d"}
CASINOSPORT_MAPPINGS = {
    "[S] 10 min sport basic wp": "10min",
    "[S] 1h sport basic wp": "1h",
    "[S] 1d 2 BLOCKS (basic wp + highroller)": "1d",
    "[S] 3d casino 1st dep total wp": "4d",
    "[S] 5d casino 1st dep": "6d",
    "[S] 7d 2 BLOCKS SPORT + CAS": "8d",
    "[S] 10d A: freebet + 100fs": "10d",
    "[S] 10d b: 150%sport + 100fs": "10d",
    "[S] 12d A: freebet + 100fs": "12d",
    "[S] 12d b: 150%sport + 100fs": "12d",
}

METRICS = ["sent", "delivered", "opened", "clicked", "converted", "unsubscribed"]
WEEK_COLUMNS = {
    'week1': 'P', 'week2': 'O', 'week3': 'N', 'week4': 'M', 'week5': 'L', 'week6': 'K', 'week7': 'J', 'week8': 'I',
    'week9': 'H', 'week10': 'G', 'week11': 'F', 'week12': 'E'
}
TIMING_BLOCKS = {
    "10min": {"casino_rows": [3, 8]},
    "1h": {"casino_rows": [9, 14]},
    "1d": {"casino_rows": [15, 20]},
    "3d": {"casino_rows": [21, 26], "section_1_rows": [93, 98], "section_2_rows": [141, 146]},
    "4d": {"casino_rows": [27, 32], "section_1_rows": [99, 104], "section_2_rows": [147, 152]},
    "6d": {"casino_rows": [33, 38], "section_1_rows": [105, 110], "section_2_rows": [153, 158]},
    "8d": {"casino_rows": [39, 44], "section_1_rows": [111, 116], "section_2_rows": [159, 164]},
    "10d": {"casino_rows": [45, 50], "section_1_rows": [117, 122], "section_2_rows": [165, 170]},
    "12d": {"casino_rows": [51, 56]},
}

WEEK_MAPPINGS = {
    'source': {
        '01': 'P', '02': 'O', '03': 'N', '04': 'M', '05': 'L', '06': 'K', '07': 'J', '08': 'I',
        '09': 'H', '10': 'G', '11': 'F', '12': 'E'
    },
    'target': {
        '01': 'BF', '02': 'BE', '03': 'BD', '04': 'BC', '05': 'BB', '06': 'BA', '07': 'AZ', '08': 'AY',
        '09': 'AX', '10': 'AW', '11': 'AV', '12': 'AU'
    }
}

SHEET_MAPPINGS = {
    'casino-ret': 'WP Chains Sport'
}

TEMPLATE_MAPPINGS = {
    "10min": "10 min",
    "1h": "1h",
    "1d": "1d",
    "3d": "3d",
    "4d": "4d",
    "6d": "6d",
    "8d": "8d",
    "10d": "10d",
    "12d": "12d"
}


@register_plugin
class CasinoRetPlugin(BaseReportPlugin):
    """Casino-Ret report plugin."""
    
    name = "casino-ret"
    supports_multiple_files = True
    
    def __init__(self):
        self.existing_excel = None
        self.replace_week = None
    
    def _detect_section_by_campaign(self, data: pd.DataFrame, file_name: str) -> Dict[str, Any]:
        """Detect section based on campaign name."""
        if data.empty or 'campaign_name' not in data.columns:
            logger.error(f"Cannot detect section for {file_name}: missing campaign_name column")
            return None
        
        campaign_name = data['campaign_name'].iloc[0]
        campaign_lower = campaign_name.lower()
        
        # Exact match first
        if campaign_name in CAMPAIGN_SECTION_MAPPINGS:
            config = CAMPAIGN_SECTION_MAPPINGS[campaign_name]
            logger.info(f"Campaign detected: '{campaign_name}' → {config['label']} (row {config['start_row']})")
            return config
        
        # Fuzzy match
        if "casino+sport" in campaign_lower or "a/b" in campaign_lower:
            config = CAMPAIGN_SECTION_MAPPINGS["casino+sport A/B Reg_No_Dep"]
            logger.info(f"Campaign detected (fuzzy): '{campaign_name}' → {config['label']} (row {config['start_row']})")
            return config
        elif "ret 1 dep" in campaign_lower:
            config = CAMPAIGN_SECTION_MAPPINGS["Ret 1 dep [SPORT] ⚽️"]
            logger.info(f"Campaign detected (fuzzy): '{campaign_name}' → {config['label']} (row {config['start_row']})")
            return config
        elif "ret 2 dep" in campaign_lower:
            config = CAMPAIGN_SECTION_MAPPINGS["Ret 2 dep [SPORT] ⚽️"]
            logger.info(f"Campaign detected (fuzzy): '{campaign_name}' → {config['label']} (row {config['start_row']})")
            return config
        
        logger.error(f"Unknown campaign: '{campaign_name}' in {file_name}. Expected: {list(CAMPAIGN_SECTION_MAPPINGS.keys())}")
        return None
    
    def process_csv(self, csv_paths: List[Path]) -> Dict[str, pd.DataFrame]:
        """Read multiple CSV files."""
        data_files = {}
        for path in csv_paths:
            df = pd.read_csv(path)
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
            data_files[path.name] = df
            logger.info(f"Loaded {path.name}: {len(df)} rows")
        return data_files
    
    def transform_data(self, data_files: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, pd.DataFrame]]:
        """Transform data for WP-Chains-2 structure."""
        report_data = {}
        
        for file_name, data in data_files.items():
            # Detect section by campaign
            section_config = self._detect_section_by_campaign(data, file_name)
            
            if not section_config:
                logger.warning(f"Skipping {file_name}: unable to detect section")
                continue
            
            # Determine template mappings
            if section_config["section"] == "casino":
                mappings = CASINOSPORT_MAPPINGS
            else:
                mappings = RETENTION_MAPPINGS
            
            # Filter templates
            filtered = data[data['template_name'].isin(mappings.keys())]
            
            # Aggregate by weeks
            weekly_data = {}
            for i, (start_date, end_date) in enumerate(WEEKLY_BOUNDARIES, 1):
                start = pd.to_datetime(start_date + ' 00:00:00')
                end = pd.to_datetime(end_date + ' 23:59:59')
                week_mask = (filtered['datetime'] >= start) & (filtered['datetime'] <= end)
                week_df = filtered[week_mask].groupby('template_name')[METRICS].sum().reset_index()
                weekly_data[f'week{i}'] = week_df
            
            # Group by timing category
            file_report = {}
            for template_name, timing_category in mappings.items():
                timing_data = {}
                for week_key, week_df in weekly_data.items():
                    template_data = week_df[week_df['template_name'] == template_name]
                    if not template_data.empty:
                        template_data = self._calculate_percentages(template_data)
                        timing_data[week_key] = template_data
                    else:
                        timing_data[week_key] = self._empty_data(template_name)
                file_report[timing_category] = timing_data
            
            report_data[file_name] = file_report
        
        return report_data
    
    def _calculate_percentages(self, data: pd.DataFrame) -> pd.DataFrame:
        """Calculate percentage metrics."""
        result = data.copy()
        result['pct_delivered'] = (result['delivered'] / result['sent'] * 100).fillna(0)
        result['pct_open'] = (result['opened'] / result['delivered'] * 100).fillna(0)
        result['pct_click'] = (result['clicked'] / result['delivered'] * 100).fillna(0)
        result['pct_cr'] = (result['converted'] / result['delivered'] * 100).fillna(0)
        return result
    
    def _empty_data(self, template_name: str) -> pd.DataFrame:
        """Create empty data row."""
        return pd.DataFrame([{
            'template_name': template_name, 'sent': 0, 'delivered': 0, 'opened': 0,
            'clicked': 0, 'converted': 0, 'unsubscribed': 0,
            'pct_delivered': 0, 'pct_open': 0, 'pct_click': 0, 'pct_cr': 0
        }])
    
    def generate_excel(self, report_data: Dict[str, Dict[str, pd.DataFrame]], output_path: Path):
        """Generate Excel file."""
        wb = Workbook()
        ws = wb.active
        
        # Headers
        week_headers = {
        'week8': ("08", "16.02"), 'week7': ("07", "09.02"),
        'week6': ("06", "02.02"), 'week5': ("05", "26.01"), 'week4': ("04", "19.01"),
        'week3': ("03", "12.01"), 'week2': ("02", "05.01"), 'week1': ("01", "29.12")
    }
        for week_key, col_letter in WEEK_COLUMNS.items():
            if week_key in week_headers:
                week_display, date = week_headers[week_key]
                ws[f'{col_letter}1'] = f"Week {week_display}\n{date}"
        
        # Section headers
        ws['A3'] = "Signed up"
        ws['A75'] = "deposits_quantity is 1"
        ws['A123'] = "deposits_quantity is 2"
        
        # Populate sections
        for file_name, section_data in report_data.items():
            # Get original data to access campaign_name
            original_data = None
            if hasattr(self, '_original_files'):
                for path in self._original_files:
                    if path.name == file_name:
                        original_data = pd.read_csv(path)
                        break
            
            # Detect section by campaign
            section_config = None
            if original_data is not None and not original_data.empty:
                section_config = self._detect_section_by_campaign(original_data, file_name)
            
            if section_config:
                self._populate_section(
                    ws,
                    section_data,
                    section_config["start_row"],
                    section_config["label"],
                    section_config["section"]
                )
            else:
                logger.warning(f"Skipping {file_name}: unable to detect section")
        
        wb.save(output_path)
        logger.info(f"Excel saved: {output_path}")
        
        # Week replacement if requested
        if self.existing_excel and self.replace_week:
            self._replace_week(output_path, self.existing_excel, self.replace_week)
    
    def _populate_section(self, ws, section_data: Dict, start_row: int, campaign_name: str, section_type: str = "retention"):
        """Populate casino or retention section."""
        ws[f'B{start_row}'] = campaign_name
        
        for timing_category, block_info in TIMING_BLOCKS.items():
            # Skip if timing category not in section data
            if timing_category not in section_data:
                continue
            
            # Get correct row range based on section type
            if section_type == "casino":
                if "casino_rows" not in block_info:
                    continue
                block_start, block_end = block_info["casino_rows"]
            elif start_row == 75:
                block_start, block_end = block_info["section_1_rows"]
            else:
                block_start, block_end = block_info["section_2_rows"]
            
            ws[f'C{block_start}'] = timing_category
            
            if timing_category in section_data:
                timing_data = section_data[timing_category]
                metrics = ["sent", "delivered", "opened", "clicked", "unsubscribed", "pct_delivered"]
                
                for i, metric in enumerate(metrics):
                    row = block_start + i
                    ws[f'D{row}'] = metric.replace('_', ' ').title()
                    
                    for week_key, col_letter in WEEK_COLUMNS.items():
                        if week_key in timing_data and not timing_data[week_key].empty:
                            df = timing_data[week_key]
                            if metric in df.columns:
                                ws[f'{col_letter}{row}'] = df[metric].iloc[0]
                        else:
                            ws[f'{col_letter}{row}'] = 0
    
    def _replace_week(self, generated_path: Path, existing_path: Path, week_number: str):
        """Replace week data in existing Excel."""
        source_col = WEEK_MAPPINGS['source'][week_number]
        target_col = WEEK_MAPPINGS['target'][week_number]
        
        generated_wb = load_workbook(generated_path, data_only=True)
        existing_wb = load_workbook(existing_path, data_only=False)
        
        generated_ws = generated_wb.active
        
        # Get target sheet based on report type
        sheet_name = SHEET_MAPPINGS.get(self.name, existing_wb.sheetnames[0])
        if sheet_name in existing_wb.sheetnames:
            existing_ws = existing_wb[sheet_name]
        else:
            existing_ws = existing_wb.active
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet")
        
        # Copy formatting
        self._copy_formatting(existing_ws, 'BE', target_col)
        
        # Copy data
        copied = 0
        current_campaign = None
        current_template = None
        
        for source_row in range(1, generated_ws.max_row + 1):
            source_b = generated_ws[f'B{source_row}'].value
            source_c = generated_ws[f'C{source_row}'].value
            source_d = generated_ws[f'D{source_row}'].value
            
            # Update current campaign if found
            if source_b and str(source_b).strip():
                current_campaign = str(source_b).strip()
            
            # Update current template if found
            if source_c and str(source_c).strip():
                current_template = str(source_c).strip()
            
            # Process metric rows
            if current_campaign and current_template and source_d:
                metric = str(source_d).strip()
                
                if metric in ["Sent", "Delivered", "Opened", "Clicked", "Unsubscribed"]:
                    target_row = self._find_target_row(existing_ws, current_campaign, current_template, metric)
                    if target_row:
                        value = generated_ws[f'{source_col}{source_row}'].value
                        if value is not None:
                            existing_ws[f'{target_col}{target_row}'].value = value
                            copied += 1
                            logger.info(f"Copied {current_campaign}/{current_template}/{metric}: {value} to row {target_row}")

        # Save the updated existing workbook back to the generated_path
        # This ensures the API returns the updated file instead of the intermediate report
        existing_wb.save(generated_path)
        logger.info(f"Updated Excel saved to output path: {generated_path} ({copied} values)")
    
    def _copy_formatting(self, ws, source_col: str, target_col: str):
        """Copy column formatting."""
        for row in range(1, ws.max_row + 1):
            source_cell = ws[f'{source_col}{row}']
            target_cell = ws[f'{target_col}{row}']
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
    
    def _find_target_row(self, ws, campaign: str, template: str, metric: str) -> int:
        """Find target row in existing Excel."""
        target_template = TEMPLATE_MAPPINGS.get(template, template)
        
        # First find the campaign section
        campaign_start = None
        for row in range(1, ws.max_row + 1):
            cell_b = ws[f'B{row}'].value
            if cell_b:
                cell_b_lower = str(cell_b).lower()
                campaign_lower = campaign.lower()
                # Bidirectional matching: either can contain the other
                if campaign_lower in cell_b_lower or cell_b_lower in campaign_lower:
                    campaign_start = row
                    break
        
        if not campaign_start:
            return None
        
        # Within campaign section, find the template timing block
        # Check both column C and D for template (different Excel formats)
        for row in range(campaign_start, min(campaign_start + 100, ws.max_row + 1)):
            cell_c = ws[f'C{row}'].value
            cell_d = ws[f'D{row}'].value
            
            # Check if template is in column C or D
            template_found = False
            if cell_c and target_template.lower() in str(cell_c).lower():
                template_found = True
            elif cell_d and target_template.lower() in str(cell_d).lower():
                template_found = True
            
            if template_found:
                # Found the template block, now find the metric row
                metric_order = ["Sent", "Delivered", "Opened", "Clicked", "Unsubscribed"]
                if metric in metric_order:
                    return row + metric_order.index(metric)
        
        return None
    
    def execute(self, input_paths: List[Path], output_path: Path, existing_excel: Path = None, replace_week: str = None):
        """Execute with optional week replacement."""
        self.existing_excel = existing_excel
        self.replace_week = replace_week
        self._original_files = input_paths  # Store for campaign detection in generate_excel
        
        for path in input_paths:
            self.validate_input(path)
        
        data_files = self.process_csv(input_paths)
        report_data = self.transform_data(data_files)
        self.generate_excel(report_data, output_path)

    def get_input_slots(self) -> List[Dict[str, Any]]:
        """Return list of expected input slots."""
        return [
            {
                "id": "casino_sport",
                "label": "Casino & Sport Data",
                "description": "Upload CSV(s) containing '[S] ...' campaigns.",
                "required": True,
                "accept": ".csv"
            },
            {
                "id": "retention",
                "label": "Retention Data",
                "description": "Upload CSV(s) containing 'Day X' campaigns.",
                "required": True,
                "accept": ".csv"
            }
        ]

    def get_expected_templates(self) -> Dict[str, List[str]]:
        """Return mapping of slot IDs to expected template names."""
        return {
            "casino_sport": list(CASINOSPORT_MAPPINGS.keys()),
            "retention": list(RETENTION_MAPPINGS.keys())
        }

    def validate_data(self, input_paths: List[Path]) -> Dict[str, Any]:
        """Analyze data and return a detailed summary of matches."""
        import difflib
        
        all_known_templates = list(CASINOSPORT_MAPPINGS.keys()) + list(RETENTION_MAPPINGS.keys())
        
        summary = {
            "total_files": len(input_paths),
            "categories": {
                "casino_sport": {"matched_rows": 0, "matched_templates": set()},
                "retention": {"matched_rows": 0, "matched_templates": set()},
                "unknown": {"rows": 0, "templates": set()}
            },
            "suggestions": {},
            "files": []
        }

        for path in input_paths:
            try:
                df = pd.read_csv(path)
                file_summary = {
                    "name": path.name,
                    "rows": len(df),
                    "matches": {"casino_sport": 0, "retention": 0, "unknown": 0}
                }

                if not df.empty and 'template_name' in df.columns:
                    for template in df['template_name'].unique():
                        template_rows = len(df[df['template_name'] == template])
                        
                        if template in CASINOSPORT_MAPPINGS:
                            file_summary["matches"]["casino_sport"] += template_rows
                            summary["categories"]["casino_sport"]["matched_rows"] += template_rows
                            summary["categories"]["casino_sport"]["matched_templates"].add(template)
                        elif template in RETENTION_MAPPINGS:
                            file_summary["matches"]["retention"] += template_rows
                            summary["categories"]["retention"]["matched_rows"] += template_rows
                            summary["categories"]["retention"]["matched_templates"].add(template)
                        else:
                            file_summary["matches"]["unknown"] += template_rows
                            summary["categories"]["unknown"]["rows"] += template_rows
                            summary["categories"]["unknown"]["templates"].add(template)
                            
                            # Add suggestion if not an exact match
                            matches = difflib.get_close_matches(template, all_known_templates, n=1, cutoff=0.6)
                            if matches:
                                summary["suggestions"][template] = matches[0]
                
                summary["files"].append(file_summary)
            except Exception as e:
                summary["files"].append({"name": path.name, "error": str(e)})

        # Convert sets to lists for JSON serialization
        summary["categories"]["casino_sport"]["matched_templates"] = list(summary["categories"]["casino_sport"]["matched_templates"])
        summary["categories"]["retention"]["matched_templates"] = list(summary["categories"]["retention"]["matched_templates"])
        summary["categories"]["unknown"]["templates"] = list(summary["categories"]["unknown"]["templates"])
        
        return summary
