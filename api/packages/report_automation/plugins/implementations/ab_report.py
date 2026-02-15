"""A-B Report Plugin - V3 Specification Implementation."""

import pandas as pd
from pathlib import Path
from typing import Dict
import logging

from ..base import BaseReportPlugin, register_plugin
from ...domain.models import ProcessedData

logger = logging.getLogger(__name__)

# V3 Constants
TEMPLATE_MAPPING = {
    "[S] 10 min sport basic wp": "10m",
    "[S] 1h sport basic wp": "1h",
    "[S] 1d 2 BLOCKS (basic wp + highroller)": "1d",
    "[S] 3d casino 1st dep total wp": "3d",
    "[S] 5d casino 1st dep": "5d",
    "[S] 7d 2 BLOCKS SPORT + CAS": "7d",
    "[S] 10d A: freebet + 100fs": "9d",
    "[S] 10d b: 150%sport + 100fs": "9d",
    "[S] 12d A: freebet + 100fs": "12d",
    "[S] 12d b: 150%sport + 100fs": "12d",
}

WEEKLY_BOUNDARIES = [
    ("2025-12-29", "2026-01-04"),  # Week 1
    ("2026-01-05", "2026-01-11"),  # Week 2
    ("2026-01-12", "2026-01-18"),  # Week 3
    ("2026-01-19", "2026-01-25"),  # Week 4
    ("2026-01-26", "2026-02-01"),  # Week 5
]

TIME_PERIODS = ["10m", "1h", "1d", "3d", "5d", "7d", "9d", "12d"]
METRICS = ["sent", "delivered", "opened", "clicked", "converted", "unsubscribed"]


@register_plugin
class ABReportPlugin(BaseReportPlugin):
    """A-B Report generation plugin."""
    
    name = "a-b-report"
    supports_multiple_files = False
    
    def process_csv(self, csv_path: Path) -> pd.DataFrame:
        """Read and process CSV file."""
        data = pd.read_csv(csv_path)
        data['datetime'] = pd.to_datetime(data['timestamp'], unit='s')
        return data
    
    def transform_data(self, data: pd.DataFrame) -> Dict[str, Dict[str, pd.DataFrame]]:
        """Transform data according to V3 specification."""
        # Filter to mapped templates only
        filtered = data[data['template_name'].isin(TEMPLATE_MAPPING.keys())]
        
        # Aggregate by weeks
        weekly_data = self._aggregate_by_weeks(filtered)
        
        # Group by time periods
        report_data = {}
        for time_period in TIME_PERIODS:
            period_templates = [t for t, p in TEMPLATE_MAPPING.items() if p == time_period]
            period_data = {}
            
            for week_key, week_df in weekly_data.items():
                week_period = week_df[week_df['template_name'].isin(period_templates)]
                if not week_period.empty:
                    aggregated = week_period[METRICS].sum()
                    period_data[week_key] = self._calculate_percentages(pd.DataFrame([aggregated]))
            
            report_data[time_period] = period_data
        
        return report_data
    
    def _aggregate_by_weeks(self, data: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Aggregate data by weekly boundaries."""
        weekly_data = {}
        for i, (start_date, end_date) in enumerate(WEEKLY_BOUNDARIES, 1):
            start = pd.to_datetime(start_date + ' 00:00:00')
            end = pd.to_datetime(end_date + ' 23:59:59')
            
            week_mask = (data['datetime'] >= start) & (data['datetime'] <= end)
            week_data = data[week_mask].groupby('template_name')[METRICS].sum().reset_index()
            weekly_data[f'week{i}'] = week_data
        
        return weekly_data
    
    def _calculate_percentages(self, data: pd.DataFrame) -> pd.DataFrame:
        """Calculate percentage metrics."""
        result = data.copy()
        result['pct_delivered'] = (result['delivered'] / result['sent'] * 100).fillna(0)
        result['pct_open'] = (result['opened'] / result['delivered'] * 100).fillna(0)
        result['pct_click'] = (result['clicked'] / result['delivered'] * 100).fillna(0)
        result['pct_cr'] = (result['converted'] / result['delivered'] * 100).fillna(0)
        return result
    
    def generate_excel(self, report_data: Dict[str, Dict[str, pd.DataFrame]], output_path: Path):
        """Generate Excel file with V3 formatting."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws.merge_cells('L1:U1')
        ws['L1'] = "Sport B"
        ws.merge_cells('L2:U2')
        ws['L2'] = "280% up to 375 EUR"
        
        # Week headers
        week_labels = ["week 1 29.12", "week 2 05.01", "week 3 12.01", "week 4 19.01", "week 5 26.01"]
        week_cols = ['I', 'L', 'O', 'R', 'U']
        for col, label in zip(week_cols, week_labels):
            ws[f'{col}3'] = label
        ws['H3'] = "Total"
        
        # Data rows
        current_row = 4
        ws[f'A{current_row}'] = "Time"
        current_row += 1
        
        metric_labels = ["Sent", "Delivered", "Opened", "Clicked", "Converted (Dep/Acc.Bon)", 
                        "Unsubscribe", "% Delivered", "% Open", "% Click", "% CR"]
        
        for time_period in TIME_PERIODS:
            ws[f'A{current_row}'] = time_period
            current_row += 1
            
            period_data = report_data.get(time_period, {})
            
            for metric_label in metric_labels:
                ws[f'A{current_row}'] = metric_label
                self._populate_metric_row(ws, current_row, metric_label, period_data)
                current_row += 1
        
        wb.save(output_path)
        logger.info(f"Excel report saved to: {output_path}")
    
    def _populate_metric_row(self, ws, row: int, metric_label: str, period_data: Dict[str, pd.DataFrame]):
        """Populate a single metric row."""
        metric_map = {
            "Sent": "sent", "Delivered": "delivered", "Opened": "opened",
            "Clicked": "clicked", "Converted (Dep/Acc.Bon)": "converted",
            "Unsubscribe": "unsubscribed", "% Delivered": "pct_delivered",
            "% Open": "pct_open", "% Click": "pct_click", "% CR": "pct_cr"
        }
        
        metric_col = metric_map.get(metric_label)
        if not metric_col:
            return
        
        total_value = 0
        week_values = []
        
        for week_key in ['week1', 'week2', 'week3', 'week4', 'week5']:
            week_df = period_data.get(week_key)
            if week_df is not None and not week_df.empty and metric_col in week_df.columns:
                value = week_df[metric_col].iloc[0]
                week_values.append(value)
                if not metric_col.startswith('pct_'):
                    total_value += value
            else:
                week_values.append(0)
        
        # Total column
        if metric_col.startswith('pct_') and week_values:
            ws[f'H{row}'] = sum(week_values) / len([v for v in week_values if v > 0]) if any(week_values) else 0
        else:
            ws[f'H{row}'] = total_value
        
        # Weekly columns
        for week_val, col in zip(week_values, ['I', 'L', 'O', 'R', 'U']):
            ws[f'{col}{row}'] = week_val
