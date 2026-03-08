"""AWOL + Casino-Ret Combined Plugin - Single-pass multi-sheet report."""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import logging
import openpyxl
import copy

from ..base import BaseReportPlugin, register_plugin
from ..constants import WEEKLY_BOUNDARIES

logger = logging.getLogger(__name__)

# AWOL template mappings
AWOL_MAPPINGS = {"Day 1": "1d", "Day 3": "3d", "Day 5": "5d", "Day 10": "10d", "Day 15": "15d", "Day 20": "20d", "Day 30": "30d", "Day 40": "40d"}

# Casino-Ret template mappings
RETENTION_MAPPINGS = {"Day 3": "3d", "Day 4": "4d", "Day 6": "6d", "Day 8": "8d", "Day 10": "10d"}
CASINOSPORT_MAPPINGS = {
    "[S] 10 min sport basic wp": "10min",
    "[S] 1h sport basic wp": "1h",
    "[S] 1d 2 BLOCKS (basic wp + highroller)": "1d",
    "[S] 3d casino 1st dep total wp": "4d",
    "[S] 5d casino 1st dep": "6d",
    "[S] 7d 2 BLOCKS SPORT + CAS": "8d",
    "[S] 10d A: freebet + 100fs": "10d",
    "[S] 10d b: 150%sport + 100fs": "10d",
    "[S] 12d A: freebet + 100fs": "12d",
    "[S] 12d b: 150%sport + 100fs": "12d",
}

CASINO_TEMPLATE_MAPPINGS = {
    "10min": "10 min", "1h": "1h", "1d": "1d", "3d": "3d", "4d": "4d",
    "6d": "6d", "8d": "8d", "10d": "10d", "12d": "12d"
}

# Week column mappings (target columns in existing Excel)
WEEK_COLUMNS = {
    'awol': {'01': 'BF', '02': 'BE', '03': 'BD', '04': 'BC', '05': 'BB', '06': 'BA', '07': 'AZ', '08': 'AY'},
    'casino-ret': {'01': 'BF', '02': 'BE', '03': 'BD', '04': 'BC', '05': 'BB', '06': 'BA', '07': 'AZ', '08': 'AY'}
}


@register_plugin
class AWOLCasinoRetPlugin(BaseReportPlugin):
    """Combined AWOL and Casino-Ret report plugin with single-pass processing."""
    
    name = "awol-casino-ret"
    supports_multiple_files = True
    
    def _detect_campaign_type(self, df: pd.DataFrame) -> Optional[Dict]:
        """Detect campaign type and return config."""
        if 'campaign_name' not in df.columns or df.empty:
            return None
        
        campaign_name = df['campaign_name'].iloc[0]
        campaign_lower = campaign_name.lower()
        
        # AWOL campaigns
        if "inactive 7" in campaign_lower:
            return {"type": "awol", "sheet": "AWOL Chains Sport", "campaign": campaign_name}
        elif "inactive 14" in campaign_lower:
            return {"type": "awol", "sheet": "AWOL Chains Sport", "campaign": campaign_name}
        elif "inactive 21" in campaign_lower or "inactive 22" in campaign_lower:
            return {"type": "awol", "sheet": "AWOL Chains Sport", "campaign": campaign_name}
        elif "inactive 30" in campaign_lower or "inactive 31" in campaign_lower:
            return {"type": "awol", "sheet": "AWOL Chains Sport", "campaign": campaign_name}
        # Casino-Ret campaigns
        elif "reg_no_dep" in campaign_lower or "casino+sport" in campaign_lower:
            return {"type": "casino-ret", "sheet": "WP Chains Sport", "campaign": campaign_name}
        elif "ret 1 dep" in campaign_lower or "retention 1" in campaign_lower:
            return {"type": "casino-ret", "sheet": "WP Chains Sport", "campaign": campaign_name}
        elif "ret 2 dep" in campaign_lower or "retention 2" in campaign_lower:
            return {"type": "casino-ret", "sheet": "WP Chains Sport", "campaign": campaign_name}
        
        return None
    
    def _calculate_metrics_awol(self, df: pd.DataFrame) -> Dict:
        """Calculate AWOL metrics (8 metrics per template)."""
        sent = df['sent'].sum()
        delivered = df['delivered'].sum()
        opened = df['opened'].sum()
        clicked = df['clicked'].sum()
        converted = df['converted'].sum() if 'converted' in df.columns else 0
        unsubscribed = df['unsubscribed'].sum()
        
        pct_delivered = (delivered / sent * 100) if sent > 0 else 0
        pct_open = (opened / delivered * 100) if delivered > 0 else 0
        pct_click = (clicked / delivered * 100) if delivered > 0 else 0
        
        return {
            'sent': sent, 'delivered': delivered, 'opened': opened, 'clicked': clicked,
            'unsubscribed': unsubscribed, 'pct_delivered': pct_delivered,
            'pct_open': pct_open, 'pct_click': pct_click
        }
    
    def _calculate_metrics_casino(self, df: pd.DataFrame) -> Dict:
        """Calculate Casino-Ret metrics (5 metrics per template)."""
        sent = df['sent'].sum()
        delivered = df['delivered'].sum()
        opened = df['opened'].sum()
        clicked = df['clicked'].sum()
        unsubscribed = df['unsubscribed'].sum()
        
        return {
            'sent': sent, 'delivered': delivered, 'opened': opened,
            'clicked': clicked, 'unsubscribed': unsubscribed
        }
    
    def _find_awol_block(self, ws, campaign: str, template: str) -> Optional[int]:
        """Find AWOL block by campaign (col B) + template (col D)."""
        campaign_start = None
        for row in range(1, ws.max_row + 1):
            cell_b = ws[f'B{row}'].value
            if cell_b and campaign.lower() in str(cell_b).lower():
                campaign_start = row
                break
        
        if not campaign_start:
            return None
        
        for row in range(campaign_start, min(campaign_start + 100, ws.max_row + 1)):
            cell_d = ws[f'D{row}'].value
            if cell_d and template.lower() in str(cell_d).lower():
                return row
            cell_b = ws[f'B{row}'].value
            if cell_b and row > campaign_start and str(cell_b).strip():
                break
        
        return None
    
    def _find_casino_row(self, ws, campaign: str, timing: str, metric: str) -> Optional[int]:
        """Find Casino-Ret row by campaign (col B) + timing (col C or D) + metric."""
        target_timing = CASINO_TEMPLATE_MAPPINGS.get(timing, timing)
        
        # First find the campaign section
        campaign_start = None
        for row in range(1, ws.max_row + 1):
            cell_b = ws[f'B{row}'].value
            if cell_b:
                cell_b_lower = str(cell_b).lower()
                campaign_lower = campaign.lower()
                # Bidirectional matching: either can contain the other
                if campaign_lower in cell_b_lower or cell_b_lower in campaign_lower:
                    campaign_start = row
                    break
        
        if not campaign_start:
            return None
        
        # Within campaign section, find the template timing block
        # Check both column C and D for template (different Excel formats)
        for row in range(campaign_start, min(campaign_start + 200, ws.max_row + 1)):
            cell_c = ws[f'C{row}'].value
            cell_d = ws[f'D{row}'].value
            
            # Check if template is in column C or D
            template_found = False
            if cell_c and target_timing.lower() in str(cell_c).lower():
                template_found = True
            elif cell_d and target_timing.lower() in str(cell_d).lower():
                template_found = True
            
            if template_found:
                # Found the template block, now find the metric row
                metric_order = ["Sent", "Delivered", "Opened", "Clicked", "Unsubscribed"]
                if metric in metric_order:
                    return row + metric_order.index(metric)
        
        return None
    
    def _replace_week_awol(self, ws, week_data: Dict, target_col: str) -> int:
        """Replace AWOL week data using AWOL logic."""
        copied = 0
        for campaign_name, templates in week_data.items():
            for template_name, metrics in templates.items():
                target_row = self._find_awol_block(ws, campaign_name, template_name)
                if target_row:
                    metric_list = ['sent', 'delivered', 'opened', 'clicked', 'unsubscribed', 'pct_delivered', 'pct_open', 'pct_click']
                    for offset, metric_key in enumerate(metric_list):
                        target_cell = ws[f'{target_col}{target_row + offset}']
                        if isinstance(target_cell.value, str) and target_cell.value.startswith('='):
                            continue
                        target_cell.value = metrics.get(metric_key, 0)
                        copied += 1
        return copied
    
    def _replace_week_casino(self, ws, week_data: Dict, target_col: str) -> int:
        """Replace Casino-Ret week data using Casino-Ret logic."""
        copied = 0
        logger.info(f"_replace_week_casino: Processing {len(week_data)} campaigns")
        for campaign_name, timings in week_data.items():
            logger.info(f"  Campaign: {campaign_name}, Timings: {list(timings.keys())}")
            for timing_category, metrics in timings.items():
                logger.info(f"    Timing: {timing_category}, Metrics: {metrics}")
                metric_list = ["Sent", "Delivered", "Opened", "Clicked", "Unsubscribed"]
                for metric in metric_list:
                    target_row = self._find_casino_row(ws, campaign_name, timing_category, metric)
                    if target_row:
                        metric_key = metric.lower()
                        ws[f'{target_col}{target_row}'].value = metrics.get(metric_key, 0)
                        copied += 1
                        logger.info(f"      Copied {metric}={metrics.get(metric_key, 0)} to row {target_row}")
                    else:
                        logger.warning(f"      Could not find row for {campaign_name}/{timing_category}/{metric}")
        return copied
    
    def execute(self, input_paths: List[Path], output_path: Path, 
                existing_excel: Path = None, replace_week: str = None):
        """Execute combined report with single-pass processing."""
        if not existing_excel or not replace_week:
            raise ValueError("awol-casino-ret requires existing_excel and replace_week")
        
        logger.info(f"=== AWOL-Casino-Ret Combined (Single-Pass) ===")
        logger.info(f"Files: {len(input_paths)}, Week: {replace_week}")
        
        week_idx = int(replace_week) - 1
        if week_idx < 0 or week_idx >= len(WEEKLY_BOUNDARIES):
            raise ValueError(f"Invalid week: {replace_week}")
        
        week_start_str, week_end_str = WEEKLY_BOUNDARIES[week_idx]
        week_start = pd.to_datetime(week_start_str + ' 00:00:00')
        week_end = pd.to_datetime(week_end_str + ' 23:59:59')
        
        # Collect data by type
        awol_data = {}  # {campaign: {template: metrics}}
        casino_data = {}  # {campaign: {timing: metrics}}
        
        for csv_path in input_paths:
            df = pd.read_csv(csv_path)
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
            df = df[(df['datetime'] >= week_start) & (df['datetime'] <= week_end)]
            
            campaign_info = self._detect_campaign_type(df)
            if not campaign_info:
                logger.warning(f"Unknown campaign: {csv_path.name}")
                continue
            
            campaign_type = campaign_info['type']
            campaign_name = campaign_info['campaign']
            
            logger.info(f"{csv_path.name} → {campaign_type} / {campaign_name}")
            
            if campaign_type == 'awol':
                # AWOL: group by template_name
                if campaign_name not in awol_data:
                    awol_data[campaign_name] = {}
                
                for template_name in df['template_name'].unique():
                    if template_name in AWOL_MAPPINGS:
                        template_df = df[df['template_name'] == template_name]
                        awol_data[campaign_name][template_name] = self._calculate_metrics_awol(template_df)
            
            elif campaign_type == 'casino-ret':
                # Casino-Ret: group by timing category
                if campaign_name not in casino_data:
                    casino_data[campaign_name] = {}
                
                # Determine template mappings
                if any(t in CASINOSPORT_MAPPINGS for t in df['template_name'].unique()):
                    mappings = CASINOSPORT_MAPPINGS
                else:
                    mappings = RETENTION_MAPPINGS
                
                logger.info(f"  Using mappings: {list(mappings.keys())[:3]}...")
                
                for template_name, timing_category in mappings.items():
                    template_df = df[df['template_name'] == template_name]
                    if not template_df.empty:
                        logger.info(f"  Found template: {template_name} → {timing_category} ({len(template_df)} rows)")
                        if timing_category not in casino_data[campaign_name]:
                            casino_data[campaign_name][timing_category] = self._calculate_metrics_casino(template_df)
                        else:
                            # Aggregate if multiple files have same timing
                            existing = casino_data[campaign_name][timing_category]
                            new_metrics = self._calculate_metrics_casino(template_df)
                            for key in existing:
                                existing[key] += new_metrics[key]
        
        # Update existing Excel
        wb = openpyxl.load_workbook(existing_excel)
        awol_col = WEEK_COLUMNS['awol'][replace_week]
        casino_col = WEEK_COLUMNS['casino-ret'][replace_week]
        
        logger.info(f"AWOL data: {len(awol_data)} campaigns")
        logger.info(f"Casino data: {len(casino_data)} campaigns")
        for camp, timings in casino_data.items():
            logger.info(f"  {camp}: {list(timings.keys())}")
        
        copied = 0
        
        # Update AWOL sheet
        if 'AWOL Chains Sport' in wb.sheetnames and awol_data:
            ws = wb['AWOL Chains Sport']
            copied += self._replace_week_awol(ws, awol_data, awol_col)
        
        # Update Casino-Ret sheet
        if 'WP Chains Sport' in wb.sheetnames and casino_data:
            ws = wb['WP Chains Sport']
            copied += self._replace_week_casino(ws, casino_data, casino_col)
        
        wb.save(existing_excel)
        wb.close()
        
        if output_path != existing_excel:
            import shutil
            shutil.copy(existing_excel, output_path)
        
        logger.info(f"=== Complete: {copied} values copied ===")
        return copied
    
    def process_csv(self, csv_path: Path) -> pd.DataFrame:
        """Not used - uses execute() directly."""
        return pd.read_csv(csv_path)
    
    def transform_data(self, data: pd.DataFrame) -> Dict:
        """Not used - uses execute() directly."""
        return {}
    
    def generate_excel(self, report_data: Dict, output_path: Path) -> None:
        """Overridden by multi-sheet version."""
        pass
