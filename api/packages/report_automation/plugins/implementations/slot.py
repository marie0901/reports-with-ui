"""
Slot Report Plugin - Handles sport campaign data with template-level granularity
Targets: WP Chains Sport and AWOL Chains Sport sheets
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from ..base import BaseReportPlugin
from ..base.registry import register_plugin
from ..constants import WEEKLY_BOUNDARIES


# Campaign to section mappings
CAMPAIGN_SECTION_MAPPINGS = {
    "Reg_No_Dep (casino/sport)": {
        "sheet": "WP Chains Sport",
        "start_row": 3,
        "has_aggregate_row": True,  # Row 3 has formulas
        "label": "Reg_No_Dep (casino/sport)"
    },
    "Retention 1 dep (sport)": {
        "sheet": "WP Chains Sport",
        "start_row": 51,
        "filter": None,  # All templates
        "label": "Retention 1 dep (sport)"
    },
    "Retention 2 dep (sport)": {
        "sheet": "WP Chains Sport",
        "start_row": 91,
        "filter": None,
        "label": "Retention 2 dep (sport)"
    },
    "Inactive 7 (sport)": {
        "sheet": "AWOL Chains Sport",
        "start_row": 3,
        "filter": None,
        "label": "Inactive 7 (sport)"
    },
    "Inactive 14 (sport)": {
        "sheet": "AWOL Chains Sport",
        "start_row": 27,
        "filter": None,
        "label": "Inactive 14 (sport)"
    },
    "Inactive 21 (sport)": {
        "sheet": "AWOL Chains Sport",
        "start_row": 51,
        "filter": None,
        "label": "Inactive 21 (sport)"
    },
    "Inactive 30+ (sport)": {
        "sheet": "AWOL Chains Sport",
        "start_row": 75,
        "filter": None,
        "label": "Inactive 30+ (sport)"
    }
}

# Week column mappings (52 weeks, F-BE, reverse chronological)
WEEK_COLUMNS = {
    '01': 'BE', '02': 'BD', '03': 'BC', '04': 'BB',
    '05': 'BA', '06': 'AZ', '07': 'AY', '08': 'AX'
}


@register_plugin
class SlotPlugin(BaseReportPlugin):
    """Plugin for slot report type with template-level granularity"""
    
    name = "slot"
    supports_multiple_files = True
    
    def _detect_campaign(self, df: pd.DataFrame) -> Optional[Dict]:
        """Detect campaign from campaign_name column"""
        if 'campaign_name' not in df.columns:
            return None
        
        campaign_name = df['campaign_name'].iloc[0]
        
        # Exact match
        if campaign_name in CAMPAIGN_SECTION_MAPPINGS:
            return CAMPAIGN_SECTION_MAPPINGS[campaign_name]
        
        # Fuzzy match
        campaign_lower = campaign_name.lower()
        if "reg_no_dep" in campaign_lower and "casino" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Reg_No_Dep (casino/sport)"]
        elif "retention 1" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Retention 1 dep (sport)"]
        elif "retention 2" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Retention 2 dep (sport)"]
        elif "inactive 7" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Inactive 7 (sport)"]
        elif "inactive 14" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Inactive 14 (sport)"]
        elif "inactive 21" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Inactive 21 (sport)"]
        elif "inactive 30" in campaign_lower:
            return CAMPAIGN_SECTION_MAPPINGS["Inactive 30+ (sport)"]
        
        return None
    
    def _detect_week_from_filename(self, filename: str) -> Optional[str]:
        """Detect week number from filename"""
        import re
        
        # Match patterns like feb15, jan26, etc.
        match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(\d{1,2})', filename.lower())
        if not match:
            return None
        
        month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        
        month = month_map[match.group(1)]
        day = int(match.group(2))
        
        # Determine year (2025 for Dec-Jan, 2026 for rest)
        year = 2025 if month == 12 else 2026
        
        file_date = datetime(year, month, day)
        
        # Find matching week
        for week_num, (start_str, end_str) in enumerate(WEEKLY_BOUNDARIES, 1):
            start_date = datetime.strptime(start_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_str, "%Y-%m-%d")
            
            if start_date <= file_date <= end_date:
                return f"{week_num:02d}"
        
        return None
    
    def _replace_week(self, existing_excel: Path, week_data: Dict, week_num: str) -> int:
        """Replace week column in existing Excel"""
        import logging
        logger = logging.getLogger(__name__)
        
        wb = openpyxl.load_workbook(existing_excel)
        target_col = WEEK_COLUMNS.get(week_num)
        
        if not target_col:
            wb.close()
            raise ValueError(f"Week {week_num} not supported")
        
        logger.info(f"Replacing week {week_num} in column {target_col}")
        values_copied = 0
        
        # Metrics to copy (skip percentage rows which have formulas)
        metrics_to_copy = ['Sent', 'Delivered', 'Open', 'Click', 'Unsub']
        
        for sheet_name in ['WP Chains Sport', 'AWOL Chains Sport']:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            sheet_data = week_data.get(sheet_name, {})
            
            for campaign_name, templates in sheet_data.items():
                campaign_info = CAMPAIGN_SECTION_MAPPINGS.get(campaign_name)
                if not campaign_info:
                    continue
                
                # Skip aggregate rows - they all have formulas
                if campaign_info.get('has_aggregate_row', False):
                    logger.info(f"Skipping aggregate row {campaign_info['start_row']} for {campaign_name}")
                
                # Write individual template rows
                for template_name, metrics in templates.items():
                    if template_name == '_aggregate':
                        continue
                    
                    # Find template row by matching Column B
                    template_row = self._find_template_row(ws, template_name, campaign_info['start_row'])
                    if template_row:
                        # Only copy first 5 metrics
                        for offset in range(5):
                            if offset < len(metrics):
                                cell = ws[f'{target_col}{template_row + offset}']
                                cell.value = metrics[offset]
                                values_copied += 1
        
        wb.save(existing_excel)
        wb.close()
        
        logger.info(f"Week replacement complete: {values_copied} values copied to {existing_excel}")
        return values_copied
    
    def _find_template_row(self, ws, template_name: str, start_row: int) -> Optional[int]:
        """Find row number for template by matching Column B"""
        # Search within reasonable range (up to 50 rows from start)
        for row in range(start_row, start_row + 50):
            cell_value = ws[f'B{row}'].value
            if cell_value and template_name in str(cell_value):
                return row
        return None
    
    def _calculate_metrics(self, df: pd.DataFrame) -> List[float]:
        """Calculate 8 metrics for template data"""
        sent = df['sent'].sum()
        delivered = df['delivered'].sum()
        opened = df['opened'].sum()
        clicked = df['clicked'].sum()
        unsubscribed = df['unsubscribed'].sum()
        
        pct_delivered = (delivered / sent * 100) if sent > 0 else 0
        pct_open = (opened / delivered * 100) if delivered > 0 else 0
        pct_click = (clicked / delivered * 100) if delivered > 0 else 0
        
        return [sent, delivered, opened, clicked, unsubscribed, 0, pct_open, pct_click]
    
    def execute(self, input_paths, output_path=None, existing_excel=None, replace_week=None):
        """Execute slot report processing"""
        # Slot plugin requires existing_excel and replace_week
        if not existing_excel or not replace_week:
            raise ValueError("Slot plugin requires existing_excel and replace_week parameters")
        
        csv_paths = input_paths if isinstance(input_paths, list) else [input_paths]
        
        # Detect week from first file if not provided
        if not replace_week:
            replace_week = self._detect_week_from_filename(csv_paths[0].name)
            if not replace_week:
                return {"error": "Could not detect week from filename"}
        
        # Process all CSV files
        week_data = {'WP Chains Sport': {}, 'AWOL Chains Sport': {}}
        
        # Get week boundaries
        week_idx = int(replace_week) - 1
        if week_idx < 0 or week_idx >= len(WEEKLY_BOUNDARIES):
            raise ValueError(f"Invalid week number: {replace_week}")
        
        week_start_str, week_end_str = WEEKLY_BOUNDARIES[week_idx]
        week_start = pd.to_datetime(week_start_str + ' 00:00:00')
        week_end = pd.to_datetime(week_end_str + ' 23:59:59')
        
        for csv_path in csv_paths:
            df = pd.read_csv(csv_path)
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
            
            # Filter by week boundaries
            df = df[(df['datetime'] >= week_start) & (df['datetime'] <= week_end)]
            
            # Detect campaign
            campaign_info = self._detect_campaign(df)
            if not campaign_info:
                continue
            
            sheet_name = campaign_info['sheet']
            campaign_name = campaign_info['label']
            
            if campaign_name not in week_data[sheet_name]:
                week_data[sheet_name][campaign_name] = {}
            
            # Calculate per-template metrics (no filtering, match exact template names)
            for template_name in df['template_name'].unique():
                template_df = df[df['template_name'] == template_name]
                template_metrics = self._calculate_metrics(template_df)
                week_data[sheet_name][campaign_name][template_name] = template_metrics
        
        # Replace week in existing Excel
        values_copied = self._replace_week(existing_excel, week_data, replace_week)
        
        # Copy result to output_path if provided
        if output_path and output_path != existing_excel:
            import shutil
            shutil.copy(existing_excel, output_path)
        
        return values_copied
    
    def process_csv(self, csv_path: Path) -> pd.DataFrame:
        """Not used - slot uses execute() directly"""
        return pd.read_csv(csv_path)
    
    def transform_data(self, data: pd.DataFrame) -> Dict:
        """Not used - slot uses execute() directly"""
        return {}
    
    def generate_excel(self, report_data: Dict, output_path: Path) -> None:
        """Not used - slot uses execute() directly"""
        pass
