"""Excel generation implementation."""

from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ...domain.interfaces import ExcelGenerator, ExcelFormatter
from ...domain.models import ProcessedData, WorksheetLayout, CellStyle


logger = logging.getLogger(__name__)


class ExcelGeneratorImpl(ExcelGenerator):
    """Implementation of Excel file generation."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.default_font = "Arial"
        self.default_font_size = 11
    
    def create_workbook(self, layout: WorksheetLayout) -> Workbook:
        """Create new Excel workbook with specified layout."""
        logger.info(f"Creating workbook with layout: {layout.name}")
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = layout.name
        
        # Apply basic formatting
        self._apply_layout_formatting(worksheet, layout)
        
        logger.info("Workbook created successfully")
        return workbook
    
    def add_worksheet(self, workbook: Workbook, layout: WorksheetLayout) -> Any:
        """Add worksheet to existing workbook."""
        logger.info(f"Adding worksheet: {layout.name}")
        
        worksheet = workbook.create_sheet(title=layout.name)
        self._apply_layout_formatting(worksheet, layout)
        
        return worksheet
    
    def populate_data(self, worksheet: Any, data: ProcessedData) -> None:
        """Populate worksheet with processed data."""
        logger.info(f"Populating data for report type: {data.report_type}")
        
        # Add headers
        self._add_headers(worksheet, data)
        
        # Add data rows
        self._add_data_rows(worksheet, data)
        
        logger.info("Data population completed")
    
    def save_workbook(self, workbook: Workbook, output_path: Path) -> None:
        """Save workbook to file."""
        logger.info(f"Saving workbook to: {output_path}")
        
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            workbook.save(output_path)
            logger.info("Workbook saved successfully")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            raise
    
    def _apply_layout_formatting(self, worksheet: Any, layout: WorksheetLayout) -> None:
        """Apply layout-specific formatting to worksheet."""
        # Set column widths
        for column_letter, width in layout.column_widths.items():
            worksheet.column_dimensions[column_letter].width = width
        
        # Freeze panes if specified
        if layout.freeze_panes:
            freeze_cell = f"{layout.freeze_panes.column}{layout.freeze_panes.row}"
            worksheet.freeze_panes = freeze_cell
        
        logger.debug(f"Applied formatting for layout: {layout.name}")
    
    def _add_headers(self, worksheet: Any, data: ProcessedData) -> None:
        """Add header rows to worksheet."""
        # Add time period headers
        col_offset = 2  # Start after row labels
        for i, period in enumerate(data.time_periods):
            cell = worksheet.cell(row=1, column=col_offset + i)
            cell.value = period.upper()
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add metric labels
        metrics = ["Sent", "Delivered", "Opened", "Clicked", "Converted", 
                  "% Delivered", "% Open", "% Click", "% CR"]
        
        for i, metric in enumerate(metrics):
            cell = worksheet.cell(row=2 + i, column=1)
            cell.value = metric
            cell.font = Font(bold=True)
    
    def _add_data_rows(self, worksheet: Any, data: ProcessedData) -> None:
        """Add data rows to worksheet."""
        metrics = ["sent", "delivered", "opened", "clicked", "converted"]
        percentage_metrics = ["% Delivered", "% Open", "% Click", "% CR"]
        
        # Add totals data
        col_offset = 2
        for i, period in enumerate(data.time_periods):
            if period in data.totals:
                period_data = data.totals[period]
                
                # Add base metrics
                for j, metric in enumerate(metrics):
                    if metric in period_data:
                        cell = worksheet.cell(row=2 + j, column=col_offset + i)
                        cell.value = period_data[metric]
                
                # Add percentage metrics
                if period in data.percentages:
                    percentages = data.percentages[period]
                    for j, metric in enumerate(percentage_metrics):
                        if metric in percentages:
                            cell = worksheet.cell(row=2 + len(metrics) + j, column=col_offset + i)
                            cell.value = f"{percentages[metric]:.2f}%"


class ExcelFormatterImpl(ExcelFormatter):
    """Implementation of Excel formatting."""
    
    def apply_cell_style(self, worksheet: Any, cell_range: str, style: Dict[str, Any]) -> None:
        """Apply styling to cell range."""
        logger.debug(f"Applying style to range: {cell_range}")
        
        # Parse cell range and apply styles
        for row in worksheet[cell_range]:
            for cell in row:
                if 'font_bold' in style and style['font_bold']:
                    cell.font = Font(bold=True)
                
                if 'background_color' in style and style['background_color']:
                    cell.fill = PatternFill(start_color=style['background_color'], 
                                          end_color=style['background_color'], 
                                          fill_type='solid')
                
                if 'alignment' in style:
                    alignment_map = {
                        'left': 'left',
                        'center': 'center', 
                        'right': 'right'
                    }
                    cell.alignment = Alignment(horizontal=alignment_map.get(style['alignment'], 'left'))
    
    def set_column_width(self, worksheet: Any, column: str, width: float) -> None:
        """Set column width."""
        worksheet.column_dimensions[column].width = width
        logger.debug(f"Set column {column} width to {width}")
    
    def add_borders(self, worksheet: Any, cell_range: str, border_style: str) -> None:
        """Add borders to cell range."""
        border_styles = {
            'thin': Side(style='thin'),
            'thick': Side(style='thick'),
            'medium': Side(style='medium')
        }
        
        side = border_styles.get(border_style, Side(style='thin'))
        border = Border(left=side, right=side, top=side, bottom=side)
        
        for row in worksheet[cell_range]:
            for cell in row:
                cell.border = border
        
        logger.debug(f"Added {border_style} borders to range: {cell_range}")
    
    def freeze_panes(self, worksheet: Any, cell: str) -> None:
        """Freeze panes at specified cell."""
        worksheet.freeze_panes = cell
        logger.debug(f"Froze panes at cell: {cell}")


class SimpleExcelGenerator:
    """Simple Excel generator for basic reports."""
    
    def __init__(self):
        """Initialize simple generator."""
        self.generator = ExcelGeneratorImpl()
        self.formatter = ExcelFormatterImpl()
    
    def create_simple_report(self, data: ProcessedData, output_path: Path) -> None:
        """Create a simple Excel report from processed data."""
        logger.info("Creating simple Excel report")
        
        # Create basic layout
        layout = WorksheetLayout(
            name="Report",
            sections=[],
            column_widths={"A": 20, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15}
        )
        
        # Create workbook
        workbook = self.generator.create_workbook(layout)
        worksheet = workbook.active
        
        # Populate data
        self.generator.populate_data(worksheet, data)
        
        # Apply basic formatting
        self.formatter.apply_cell_style(worksheet, "A1:F1", {
            'font_bold': True,
            'background_color': 'D3D3D3',
            'alignment': 'center'
        })
        
        # Save workbook
        self.generator.save_workbook(workbook, output_path)
        
        logger.info(f"Simple report created: {output_path}")
    
    def create_basic_workbook(self, output_path: Path) -> None:
        """Create a basic workbook for testing."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Report"
        
        # Add test data
        worksheet['A1'] = "Metric"
        worksheet['B1'] = "Value"
        worksheet['A2'] = "Test Sent"
        worksheet['B2'] = 100
        worksheet['A3'] = "Test Delivered"
        worksheet['B3'] = 95
        
        # Apply formatting
        worksheet['A1'].font = Font(bold=True)
        worksheet['B1'].font = Font(bold=True)
        
        self.generator.save_workbook(workbook, output_path)
        logger.info(f"Basic workbook created: {output_path}")
