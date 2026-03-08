import sys
import os
import shutil
import tempfile
import logging
import json
import pandas as pd
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware

# Add the packages directory to the system path
sys.path.append(str(Path(__file__).parent / "packages"))

from report_automation.plugins import get_plugin, list_plugins
from report_automation.domain.models import ProcessedData

app = FastAPI()

# Configure CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.get("/api/reports")
def get_reports():
    """List available report types."""
    return {"reports": list_plugins()}

@app.get("/api/reports/{report_type}/config")
def get_report_config(report_type: str):
    """Get configuration for a specific report type."""
    plugin_class = get_plugin(report_type)
    if not plugin_class:
        raise HTTPException(status_code=404, detail=f"Report type '{report_type}' not found")
    
    plugin = plugin_class()
    return plugin.get_config()

@app.post("/api/validate")
async def validate_data(
    report_type: str = Form(...),
    files: List[UploadFile] = File(...),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """Analyze data before report generation."""
    temp_dir = Path(tempfile.mkdtemp())
    try:
        input_paths = []
        for file in files:
            file_path = temp_dir / file.filename
            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
            input_paths.append(file_path)
        
        plugin_class = get_plugin(report_type)
        if not plugin_class:
            raise HTTPException(status_code=400, detail=f"Report type '{report_type}' not found")
        
        plugin = plugin_class()
        summary = plugin.validate_data(input_paths)
        
        # Clean up temp dir in background
        background_tasks.add_task(shutil.rmtree, temp_dir, ignore_errors=True)
        return summary
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Validation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/generate")
async def generate_report(
    report_type: str = Form(...),
    files: List[UploadFile] = File(...),
    existing_excel: Optional[UploadFile] = File(None),
    replace_week: Optional[str] = Form(None),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """Generate a report from uploaded CSV files."""
    temp_dir = Path(tempfile.mkdtemp())
    output_path = temp_dir / f"{report_type}_report.xlsx"
    
    try:
        # Save uploaded CSV files
        input_paths = []
        for file in files:
            file_path = temp_dir / file.filename
            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
            input_paths.append(file_path)
            
        # Handle existing Excel file if provided
        existing_excel_path = None
        if existing_excel:
            existing_excel_path = temp_dir / existing_excel.filename
            with open(existing_excel_path, "wb") as f:
                shutil.copyfileobj(existing_excel.file, f)

        # Get plugin
        try:
            plugin_class = get_plugin(report_type)
            if not plugin_class:
                raise HTTPException(status_code=400, detail=f"Report type '{report_type}' not found")
            plugin = plugin_class()
        except ImportError as e:
             # Fallback if plugins are not registered correctly due to import issues
             # This might happen if the plugin registration relies on side-effects of imports
             # For now, let's assume standard import works, otherwise we might need to manually register
             logger.error(f"Plugin import error: {e}")
             raise HTTPException(status_code=500, detail=f"Plugin system error: {str(e)}")

        # Execute plugin logic
        try:
            # Always pass list of paths for consistency
            plugin.execute(input_paths, output_path, existing_excel_path, replace_week)
                 
        except Exception as e:
            logger.error(f"Error executing plugin: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=str(e))

        if not output_path.exists():
            raise HTTPException(status_code=500, detail="Report generation failed to create output file")

        # Return file response
        # We use BackgroundTasks to clean up the temp directory after the response is sent
        background_tasks.add_task(shutil.rmtree, temp_dir, ignore_errors=True)
        
        return FileResponse(
            path=output_path, 
            filename=output_path.name, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except HTTPException:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Unexpected error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/analyze")
async def analyze_files(
    files: List[UploadFile] = File(...),
    excel_file: Optional[UploadFile] = File(None),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """Analyze CSV and Excel files for mapping suggestions."""
    temp_dir = Path(tempfile.mkdtemp())
    try:
        csv_analysis = []
        for file in files:
            file_path = temp_dir / file.filename
            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
            
            df = pd.read_csv(file_path)
            campaigns = df['campaign_name'].unique().tolist() if 'campaign_name' in df.columns else []
            templates = df['template_name'].unique().tolist() if 'template_name' in df.columns else []
            
            csv_analysis.append({
                "filename": file.filename,
                "campaigns": campaigns,
                "templates": templates,
                "row_count": len(df)
            })
            
        excel_analysis = {"sheets": []}
        if excel_file:
            excel_path = temp_dir / excel_file.filename
            with open(excel_path, "wb") as f:
                shutil.copyfileobj(excel_file.file, f)
            
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Column B values (usually campaign or template names)
                col_b_values = []
                for row in range(1, min(ws.max_row, 500) + 1):
                    val = ws[f'B{row}'].value
                    if val and str(val).strip():
                        col_b_values.append({
                            "row": row,
                            "value": str(val).strip()
                        })
                
                excel_analysis["sheets"].append({
                    "name": sheet_name,
                    "column_b": col_b_values
                })

        background_tasks.add_task(shutil.rmtree, temp_dir, ignore_errors=True)
        return {
            "csv_analysis": csv_analysis,
            "excel_analysis": excel_analysis
        }
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Analysis error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/analyze-sheet")
async def analyze_sheet(
    excel_file: UploadFile = File(...),
    sheet_name: str = Form(...),
    column_letter: str = Form("B"),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """Analyze a specific sheet and column for labels."""
    temp_dir = Path(tempfile.mkdtemp())
    try:
        excel_path = temp_dir / excel_file.filename
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excel_file.file, f)
        
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]
        labels = []
        for row in range(1, min(ws.max_row, 1000) + 1):
            val = ws[f'{column_letter}{row}'].value
            if val and str(val).strip():
                labels.append({
                    "row": row,
                    "value": str(val).strip()
                })
        
        background_tasks.add_task(shutil.rmtree, temp_dir, ignore_errors=True)
        return {"labels": labels}
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Sheet analysis error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/report-types")
async def create_report_type(config: str = Form(...)):
    """Save a new report type configuration."""
    try:
        data = json.loads(config)
    except Exception as e:
        logger.error(f"Failed to parse config JSON: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid JSON configuration: {str(e)}")
        
    report_id = data.get('id')
    if not report_id:
        raise HTTPException(status_code=400, detail="Report ID is required")
        
    # Standardize id to be filename-safe
    report_id = report_id.lower().replace(" ", "_")
    data['id'] = report_id
    
    config_dir = Path(__file__).parent / "storage" / "configs"
    config_dir.mkdir(parents=True, exist_ok=True)
    
    config_path = config_dir / f"{report_id}.json"
    with open(config_path, "w") as f:
        json.dump(data, f, indent=2)
        
    return {"status": "success", "id": report_id}

@app.get("/api/report-types")
def list_report_configs():
    """List all custom report configurations."""
    config_dir = Path(__file__).parent / "storage" / "configs"
    if not config_dir.exists():
        return {"configs": []}
        
    configs = []
    for f in config_dir.glob("*.json"):
        with open(f, "r") as config_file:
            try:
                data = json.load(config_file)
                configs.append({
                    "id": data.get("id"),
                    "name": data.get("name", f.stem),
                    "description": data.get("description", "")
                })
            except:
                continue
    return {"configs": configs}

@app.get("/api/report-types/{report_id}")
def get_report_config_json(report_id: str):
    """Get a specific report type configuration."""
    config_path = Path(__file__).parent / "storage" / "configs" / f"{report_id}.json"
    if not config_path.exists():
        raise HTTPException(status_code=404, detail="Config not found")
        
    with open(config_path, "r") as f:
        return json.load(f)

@app.delete("/api/report-types/{report_id}")
def delete_report_type(report_id: str):
    """Delete a report type configuration."""
    config_path = Path(__file__).parent / "storage" / "configs" / f"{report_id}.json"
    if config_path.exists():
        os.remove(config_path)
        return {"status": "success"}
    raise HTTPException(status_code=404, detail="Config not found")

